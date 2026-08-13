"""Python_Essentials

### Defining Variables with Different Data Types
"""

age = 22
salary = 5.5
is_active = True

print(f"Age: {age}")
print(f"Salary: {salary}")
print(f"Is Active: {is_active}")

"""### Checking the Data Types

### Understanding Collections: List, Tuple, Dict, and Set
"""

# List: Ordered and Mutable
my_list = ["apple", "banana", "cherry", "apple"]

# Tuple: Ordered and Immutable
my_tuple = ("north", "south", "east", "west")

# Set: Unordered and Unique items only
my_set = {1, 2, 2, 3, 4, 4}

# Dictionary: Key-Value pairs
my_dict = {"brand": "Ford", "model": "Mustang", "year": 1964}

print(f"List: {my_list}")
print(f"Tuple: {my_tuple}")
print(f"Set: {my_set}")
print(f"Dict: {my_dict}")

"""### 📝 Collection Practice Questions

1. **List Modification**: Create a list of 3 colors. Add a 4th color to it using `.append()`.
2. **Tuple Access**: Create a tuple with your birth year and birth month. Try to change the month value—what happens?
3. **Set Uniqueness**: Create a list with numbers `[1, 2, 2, 3, 3, 3]`. Convert it into a `set`.
4. **Dictionary Lookup**: Create a dictionary for a 'Book' with keys: `title`, `author`, and `pages`. Print only the `author`.

### Converting Between Lists and Sets

You can use `set()` to convert a list into a set, and `list()` to convert a set back into a list.
"""

# 1. List to Set (Removing duplicates)
original_list = [1, 2, 2, 3, 3, 3, 4]
converted_set = set(original_list)
print(f"Original List: {original_list}")
print(f"Converted Set: {converted_set} (Notice duplicates are gone!)")

# 2. Set to List
back_to_list = list(converted_set)
print(f"Converted back to List: {back_to_list}")

# Tip: This is the fastest way to get unique values from a list!

"""### More Type Casting Examples
You can use the type name as a function to convert values.
"""

# String to Integer
number_str = "42"
number_int = int(number_str)
print(f"String '{number_str}' becomes int: {number_int}")

# Integer to Float
pi_int = 3
pi_float = float(pi_int)
print(f"Int {pi_int} becomes float: {pi_float}")

# Anything to String
boolean_val = True
boolean_str = str(boolean_val)
print(f"Bool {boolean_val} becomes string: '{boolean_str}'")

"""### The `tuple()` Function
Tuples also follow the same type-casting rules. You can convert a list or even a string into a tuple.
"""

# 1. List to Tuple
my_colors = ["red", "green", "blue"]
locked_colors = tuple(my_colors)
print(f"List converted to Tuple: {locked_colors}")

# 2. String to Tuple (splits it into characters)
word = "Python"
word_tuple = tuple(word)
print(f"String converted to Tuple: {word_tuple}")

"""### Converting Tuples to Other Types
You can convert a tuple into a list (to make it mutable) or a set (to remove duplicates).
"""

# 1. Tuple to List
my_tuple = (10, 20, 30)
my_list = list(my_tuple)
my_list.append(40)
print(f"Tuple {my_tuple} converted to list: {my_list}")

# 2. Tuple to Set
duplicate_tuple = (1, 1, 2, 2, 3)
my_set = set(duplicate_tuple)
print(f"Tuple {duplicate_tuple} converted to set: {my_set}")

# 3. Tuple to String
char_tuple = ('P', 'y', 't', 'h', 'o', 'n')
# Note: str(tuple) just gives a string with parentheses,
# but we often use .join() to make a clean string
word = "".join(char_tuple)
print(f"Tuple {char_tuple} joined into string: '{word}'")

"""### Working with Dictionaries
Dictionaries use curly braces `{}` and contain pairs of keys and values.
"""

# 1. Creating a Dictionary
user_profile = {
    "name": "Alice",
    "age": 25,
    "is_pro": True
}

# 2. Accessing a value using its key
print(f"User Name: {user_profile['name']}")

# 3. Adding/Modifying via square brackets (Standard Python)
user_profile["email"] = "alice@example.com"
user_profile["age"] = 26

# 4. Attempting Dot Notation (This will cause an error)
print("--- Testing Dot Notation ---")
try:
    print(user_profile.email)
except AttributeError as e:
    print(f"Error: {e}")
    print("Note: Python dictionaries require square brackets: user_profile['email']")

print(f"\nUpdated Profile: {user_profile}")

"""### 🚀 Your Turn!
You can use the empty cell below to try the **Collection Practice Questions** listed earlier.

Specifically, try Question 4: **Create a dictionary for a 'Book' with keys: `title`, `author`, and `pages`. Print only the `author`.**
"""

book = {
    "title":"Prince of Persia",
    "author": "Rayala",
    "pages":"812"
}

print(f"Book Author is: {book['author']}")

"""## Control Flow: Conditionals (`if`, `elif`, `else`)
Conditionals allow you to execute code only if a certain condition is met.

**Note:** Python uses a colon `:` and indentation to group code blocks.
"""

age = 18

if age >= 21:
    print("You are an adult and can enter the club.")
elif age >= 18:
    print("You are an adult but cannot enter the club.")
else:
    print("You are a minor.")

"""### Comparison Operators
To build conditions, we use these operators:
- `==` (Equal to)
- `!=` (Not equal to)
- `>` , `<` (Greater than, Less than)
- `>=` , `<=` (Greater/Less than or equal to)

### Logical Operators
You can combine conditions using:
- `and`: Both must be true.
- `or`: At least one must be true.
- `not`: Reverses the result.
"""

# Case 1: Has ticket and NOT banned
has_ticket = True
is_banned = False

print("--- Scenario 1: Not Banned ---")
if has_ticket and not is_banned:
    print("Access granted!")
else:
    print("Access denied.")

# Case 2: Has ticket but IS banned
is_banned = True

print("\n--- Scenario 2: Is Banned ---")
if has_ticket and not is_banned:
    print("Access granted!")
else:
    print("Access denied.") # This runs because 'not True' is 'False'

"""### 🚀 Conditionals Practice
Try this: Create a variable `score = 85`. Write an `if/elif/else` chain that prints "A" for scores 90+, "B" for 80-89, and "C" for anything else.
"""

score = 91

if score >= 90:
  print("A")
elif score >= 80 and score <= 89:
  print("B")
else:
  print("C")

"""## Control Flow: Loops
Loops allow you to repeat a block of code multiple times.

### 1. The `for` Loop
This is most commonly used to go through items in a list or a range of numbers.
"""

fruits = ["apple", "banana", "cherry"]

print("Iterating through a list:")
for fruit in fruits:
    print(f"I like {fruit}")

print("\nUsing range(5) to count:")
for i in range(5):
    print(f"Number: {i}")

"""### 2. The `while` Loop
This runs as long as the condition remains `True`. Be careful not to create an infinite loop!
"""

count = 1
while count <= 3:
    print(f"While loop count: {count}")
    count += 1  # Increment to avoid infinite loop

"""### 🚀 Loops Practice
Try this: Create a list of 5 numbers. Write a `for` loop that calculates the sum of all numbers in that list and prints the final total.
"""

nums = [1,2,3,4,5]

sum = 0

for num in nums:
  sum += num

print(sum)

"""### 🚀 While Loop Practice
**Challenge:** Create a variable `countdown = 5`. Write a `while` loop that prints the value of `countdown` and then subtracts 1 from it each time. The loop should stop once it reaches 0 and print "Blast off!".
"""

# Write your code here!
countdown = 5

while countdown >= 0:
  if countdown == 0:
    print("Blast off !")
    break;
  print(countdown)
  countdown -= 1

"""### 💡 Bonus Concept: Loop Control (`break` and `continue`)
Sometimes you need to stop a loop early or skip an iteration:
- `break`: Stops the loop completely.
- `continue`: Skips the rest of the current block and moves to the next iteration.
"""

print("Example of break:")
for n in range(1, 10):
    if n == 5:
        break # Stops when we hit 5
    print(n)

print("\nExample of continue (skipping even numbers):")
for n in range(1, 6):
    if n % 2 == 0:
        continue # Skips the rest of the loop for even numbers
    print(n)

"""## 🔍 Deep Dive: Hidden But Important Concepts
Let's cover `range()`, math operators, and membership checks.
"""

# 1. range(start, stop, step)
print("Range with step 2:")
for n in range(0, 10, 2):
    print(n) # 0, 2, 4, 6, 8

# 2. Floor Division vs Regular Division
print(f"\nRegular: 5 / 2 = {5/2}")
print(f"Floor: 5 // 2 = {5//2}")
print(f"Remainder: 5 % 2 = {5%2}")

# 3. Membership Operators (Works on lists, strings, dicts)
fruits = ["apple", "mango"]
print(f"\nIs 'apple' in list? {'apple' in fruits}")
print(f"Is 'pear' not in list? {'pear' not in fruits}")

"""### 🚀 Quick Challenge
Can you write a `for` loop using `range()` that prints the **even** numbers between 10 and 20 (inclusive)?
"""

# Your way (Correct!)
print("Method 1 (using if):")
for a in range(10, 21):
    if a % 2 == 0:
        print(a)

# The 'Step' way (More efficient)
print("\nMethod 2 (using range step):")
for a in range(10, 21, 2):
    print(a)

"""## 🛠️ Functions: The Building Blocks
A function is a reusable block of code that only runs when it is called.

### 1. The Anatomy of a Function
- `def`: The keyword used to **def**ine a function.
- **Parameters**: The inputs inside the parentheses `()`.
- **Indentation**: The code belonging to the function must be indented.
- `return`: Sends a value back to the caller.
"""

# A function that returns a value
def add_numbers(a, b):
    return a + b

# A function that just prints (returns None by default)
def shout_hello(name):
    print(f"HELLO {name.upper()}!")

# Using the return value
result = add_numbers(5, 10)
print(f"The sum is: {result}")

# Using the print-only function
shout_hello("alice")

"""### 2. Why `return` Matters (Function Chaining)
Because `add_numbers` **returns** a value, we can use its output as the input for another function. This is what makes complex apps possible!
"""

def multiply_by_two(n):
    return n * 2

# Chaining: add_numbers -> multiply_by_two
final_value = multiply_by_two(add_numbers(5, 5))
print(f"(5 + 5) * 2 = {final_value}")

"""### 🚀 Function Challenge
Try this: Write two functions:
1. `square(n)`: Returns the square of a number.
2. `add_five(n)`: Returns a number plus 5.

Then, in one line of code, use both functions to calculate the square of 3, add 5 to that result, and print it. (The answer should be 14).
"""

# Write your functions and the chain here!
def square(n):
  return n * n

def add_five(n):
  return n + 5

result = add_five(square(3))

print(result)

"""### 🧩 Parameters vs. Arguments & Default Values

- **Parameters**: The names `name` and `message` in `def greet(name, message):`.
- **Arguments**: The actual strings `"Rayala"` and `"Good morning"` used in `greet("Rayala", "Good morning")`.

**Default Parameters** allow you to make certain inputs optional.
"""

# 'message' has a default value of "Hello"
def greet(name, message="Hello"):
    return f"{message}, {name}!"

# Call 1: Using the default value
print(greet("Alice"))

# Call 2: Overriding the default value
print(greet("Bob", "Welcome"))

"""## 🛡️ Exception Handling: Making Code Robust

Sometimes code encounters an error (an **Exception**). If we don't 'catch' it, the program stops.

### 1. The `try...except` Block
This allows you to 'try' a piece of code and provide a safety net if a specific error occurs.
"""

def divide_numbers(a, b):
    try:
        result = a / b
        return result
    except ZeroDivisionError:
        return "Error: You cannot divide by zero!"
    except TypeError:
        return "Error: Please provide numbers only."

print(f"Normal: {divide_numbers(10, 2)}")
print(f"Zero Case: {divide_numbers(10, 0)}")
print(f"Type Case: {divide_numbers(10, '5')}")

"""### 2. The `raise` Keyword
Sometimes, you want to *manually* trigger an error if a certain condition isn't met (e.g., a user enters a negative age).
"""

def check_age(age):
    if age < 0:
        raise ValueError("Age cannot be negative!")
    return f"Age {age} is valid."

try:
    print(check_age(-5))
except ValueError as e:
    print(f"Caught a custom error: {e}")

"""### 🏦 Real-World Use Case: Automated Loan Processor

Imagine we are processing a list of loan applications. Some might have missing info, some might have typos, and some might be fraudulent. We need to handle each case without crashing the whole bank's system.
"""

def process_loan(app):
    try:
        # 1. Check for missing data (KeyError)
        name = app['name']
        income = app['income']

        # 2. Check for logical errors (Manual Raise)
        if income < 0:
            raise ValueError(f"Invalid income for {name}: Cannot be negative.")

        # 3. Check for calculation errors (TypeError)
        # If income is a string like "high", this math will fail
        if isinstance(income, (int, float)):
            tax_bracket = income * 0.2
        else:
            raise TypeError("Income must be a number")

        return f"✅ Application for {name} processed. Tax estimate: {tax_bracket}"

    except KeyError as e:
        return f"❌ Data Error: Missing field {e}"
    except ValueError as e:
        return f"⚠️ Logic Error: {e}"
    except TypeError:
        return "🚫 Type Error: Income must be a number."
    except Exception as e:
        return f"❓ Unexpected Error: {e}"
    finally:
        print("System check complete for this user")

# Let's test different 'bad' data samples
apps = [
    {"name": "Alice", "income": 50000},          # Valid
    {"name": "Bob"},                             # Missing 'income'
    {"name": "Charlie", "income": -100},         # Negative income
    {"name": "David", "income": "high"}          # String instead of number
]

for a in apps:
    print(process_loan(a))

"""## 📂 Automation Fundamentals: File & Path Handling

In automation, your script needs to know *where* it is and *where* the files it needs are located. We use two main libraries:
1. `os`: The classic way to interact with the operating system.
2. `pathlib`: The modern, 'Pythonic' way to handle paths as objects.
"""

import os
from pathlib import Path

# Let's see where we are right now
current_dir = os.getcwd()
print(f"Current Working Directory: {current_dir}")

"""### 🚀 Path Challenge

Using the `Path` object imported above, can you try to define a path for a folder named `downloads`?

**Hint:** You can create a Path object like this: `my_path = Path("folder_name")`.

Once you define it, how would you check if that folder actually exists on the system? (Try looking for a method like `.exists()` on your path variable).
"""

# 1. Create a Path object for a folder named 'automation_reports'
report_folder = Path("automation_reports")

# 2. Use report_folder.mkdir(exist_ok=True) to create the folder
report_folder.mkdir(exist_ok=True)

report_folder.exists()
# (exist_ok=True prevents an error if the folder is already there)

# 3. Use an 'if' statement with report_folder.exists() to print a success message
if report_folder.exists():
  print("Yes")
else:
  print("No")

"""### 🔍 Inspecting your new folder
We can use the `/` operator to create a path to a file *inside* our new folder. In automation, this is how you define where to save your results.
"""

# 1. Create a path for a file INSIDE the folder
file_in_folder = report_folder / "daily_log.txt"
print(f"Target file path: {file_in_folder}")



# 2. Let's list everything in our current directory to see the folder
print("\nItems in current directory:")
for item in Path(".").iterdir():
    if item.is_dir():
        print(f"[Folder] {item.name}")
    else:
        print(f"[File]   {item.name}")

# 1. Use your 'file_in_folder' variable and the .write_text() method
# to write "Automation Started!" into the file.
file_in_folder.write_text("Automation Started!")


# 2. To verify it worked, let's read it back.
# Can you guess the method name to read the text? (Hint: it's the opposite of write_text)
# print(file_in_folder.read_text())
file_in_folder.exists()

file_in_folder.name

"""### 🧩 Deconstructing Paths

Automations often need to rename files or move them based on their type. Let's see how Python breaks down a path.
"""

sample_path = Path("downloads/reports/2023_financials.xlsx")

# Try to print the following using the attributes mentioned above:
# 1. The full name of the file
sample_path.name
# 2. Only the extension (suffix)
sample_path.suffix
# 3. The name without the extension (stem)
sample_path.stem

"""### 🚀 Next Step: Environment Variables

In your context, you mentioned working with folders like `data/` and `reports/`.

One common task in automation is getting the **Home Directory** of the user so the script works on any computer (Mac, Windows, or Linux).

Can you find a method in `Path` that identifies the user's home directory? (Hint: try typing `Path.` and seeing if a 'home' related method appears).
"""

# Try to find and print the Home Directory here
Path.home()

"""### 🛠️ Setting up the Project Structure

Automations usually start by ensuring the necessary folders exist. Let's practice creating the `data` folder safely.
"""

# Define a Path for a folder named 'data'
# Then use .mkdir() with the correct argument to create it safely
folder_data = Path("data")

folder_data.mkdir(exist_ok=True)

folder_data.exists()

# --- The Automation Archiver Challenge ---

# 1. Define a path for a folder named 'archive' inside your 'data' folder.
# Hint: Use the '/' operator with your existing folder_data variable.
archive_folder = folder_data/("archive")
# Create parent folders if not exists
archive_folder.mkdir(parents=True ,exist_ok=True)

if archive_folder.exists:
  print("OK. Created")
else:
  print("Error in creation")

# 3. Create a file named 'session_log.txt' inside that new archive folder.
session_file = archive_folder / "session_log.txt"
# Write the text "Log Entry: System Check Passed" into it.
session_file.write_text("Log Entry: System Check Passed")

# 4. Final Step: Print the absolute path of your new log file to the console.
# (Look for a method on your path object that sounds like 'absolute')
print(session_file.absolute())

"""### 🔍 Batch Processing & Discovery

Automation is powerful because it can handle 1 or 1,000 files using the same logic. Let's learn how to 'scan' a directory.
"""

# 1. Create a second file in your archive folder called 'old_report.pdf'
(archive_folder / "old_report.pdf").touch()

# 2. Use a 'for' loop and archive_folder.iterdir() to print every item found.
for item in Path(archive_folder).iterdir():
  if item.suffix == ".txt":
    print(f"Name: {item.name}")
  else:
    print(f"Not a text file: {item.name}")

# 3. Inside the loop, use an 'if' statement with .suffix to only print
# the names of files that end in '.txt'.

"""### 🚀 The 'Glob' Power Move

If you only care about specific files, you can use `.glob("*.extension")`. The `*` is a wildcard that means 'anything'.

How would you use `archive_folder.glob("*.txt")` to get a list of just the text files and print their count?
"""

# Try using .glob() here to find all .txt files and print how many you found.
archive_folder.glob("*.txt")

"""### Understanding Generators & Glob

When you use `.glob()`, Python returns a **generator**.

**Why?** Efficiency! Generators don't store all the results in memory. They 'generate' each item only when you ask for it (like in a loop).

To see the contents of a generator immediately, you can wrap it in `list()`:
```python
files = list(archive_folder.glob("*.txt"))
print(files)
```
"""

# 1. Use list() around your glob call to see all matching files at once
txt_files = list(archive_folder.glob("*.txt"))

print(txt_files)
len(txt_files)

"""## 🧱 Working with JSON Data

JSON is essentially a way to represent data as text. In Python, we use the built-in `json` module to convert between JSON text and Python dictionaries/lists.
"""

import json

# Imagine this string came from a web API
user_json = '{"user_id": 101, "username": "automation_pro", "active": true}'

# 1. Use json.loads() to convert the string into a Python dictionary
user_data = json.loads(user_json)

# 2. Print the type of your new 'data' variable to see what it became
print(f"Type of: {type(user_data)}")
# 3. Try to print only the 'username' from that data
print(f"Username: {user_data['username']}")

"""### Saving Data to a File

Sometimes we want to do the opposite: take a Python dictionary and save it into a `.json` file so another program can read it later. We use `json.dump()` for this.

Based on what we learned about `pathlib` earlier, how would you open a file for writing before 'dumping' your data into it?
"""

import json

# 1. We'll create a new JSON file in your archive folder
status_file = archive_folder / "status.json"

# 2. Use the 'with' syntax to open the file for writing ('w')
with open(status_file, "w") as f:
    # 3. We save the data. Notice the two arguments inside the parentheses:
    json.dump(user_data, f)

with open(status_file, "r") as f:
    loaded_data = json.load(f)
    print(loaded_data)
# Now, let's try to read it back!
# Can you complete the 'with' block below to LOAD the data from the file?
# Hint: Use "r" for read mode and json.load(f)

# with open(status_file, "r") as f:
#     loaded_data =
#     print(loaded_data)

"""### 🔍 Verifying the Raw File Content

To see the difference between a Python **Dictionary** and the actual **JSON File**, let's read the file as plain text.
"""

# Use the .read_text() method on your 'status_file' variable
# and print the result.
status_file.read_text()

"""### 📜 Introduction to Logging

In automation, logs provide a historical record of every action. Unlike `print()`, logs can be saved to files automatically and categorized by importance (e.g., Info, Warning, Error).
"""

import logging
from pathlib import Path

for handler in logging.root.handlers[:]:
    logging.root.removeHandler(handler)

reports_path = Path("automation_reports")

reports_path.mkdir(parents=True, exist_ok=True)

# Define where our log file will live
log_path = reports_path / "send_email_logs.log"

# # Setting up the logger
# # Challenge: Look at the 'level' argument below.
# # What do you think happens if we set it to logging.ERROR instead of logging.INFO?
logging.basicConfig(
    filename=log_path,
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    filemode='w',
    encoding='utf-8'
)

email = "ayush@mail.com"

logging.info(f"Email was sento to {email}")
logging.error("Email send failed")

with open(log_path, "r") as f:
  print(f.read())

# # Try adding your own log messages below using:
# # logging.info("Your message")
# # logging.error("An error message")

"""### 🚀 Your Challenge

1. Run the cell above to initialize the logger.
2. In the empty cell below, write three log messages: one `info`, one `warning`, and one `error`.
3. After running them, how would you check the contents of `robot.log` to see if your trail was recorded? (Hint: Remember the `.read_text()` method we used on JSON files!)
"""

# 1. Write your log messages here using logging.info(), logging.warning(), and logging.error()
logging.info("Info Message")
logging.warning("Warning Message")
logging.error("Error Message")

print(log_path.read_text())
# 2. Now, use log_path.read_text() to see what was written to the file.
# Remember to wrap it in a print() so you can see the output!

"""## 📊 Excel Automation with `openpyxl`

Before we can manipulate cells, we need an Excel file. Let's create one programmatically so we have a consistent starting point.
"""

import pandas as pd

# Creating a sample dataset
data = {
    "ID": ["APP001", "APP002", "APP003"],
    "Name": ["Rahul", "Priya", "Sneha"],
    "Email": ["rahul@gmail.com", "priya@gmail.com", "sneha@gmail.com"],
    "Status": ["Pending", "Pending", "Approved"]
}

# Saving it as an Excel file for our practice
df = pd.DataFrame(data)
df.to_excel("applications.xlsx", index=False, sheet_name="Applications")

print("File 'applications.xlsx' has been created with a sheet named 'Applications'.")

"""### 1. Loading the Workbook and Accessing the Sheet

Use this cell to practice loading your Excel file and targeting the specific worksheet.
"""

# 1. Import load_workbook
from openpyxl import load_workbook
# 2. Load the workbook 'applications.xlsx'
wb = load_workbook("applications.xlsx")
# 3. Access the sheet named 'Applications'
sheet = wb["Applications"]

# 1. Read and print the value of cell B2
name = sheet["B2"].value
print(name)

# 2. Write data
sheet["B2"] = "Jaadu"

wb.save("applications.xlsx")

"""### 2. Iterating Through Rows

To process all applications, we loop through the rows. Our goal is to see each row as a collection of values.
"""

# Using iter_rows to skip the header and get data
# min_row=2 starts us at the first data entry (Rahul)

for row in sheet.iter_rows(min_row=2, values_only=True):
    # row is now a tuple like ('APP001', 'Rahul', ...)
    print(f"Processing row: {row}")

    # Challenge: How would you access just the 'Email' from this row tuple?
    # (Remember that indexing starts at 0)
    # email = row[2]
    # print(email)

