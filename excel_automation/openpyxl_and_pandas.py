"""openpyxl_and_pandas.py

Continuing your journey toward becoming an Intelligent Automation Engineer.

We pick up right where python_essentials.py left off:
  → You can read/write a single cell in Excel
  → You can loop through rows

Now we go deeper — step by step.
"""

# ============================================================
# SECTION 1: openpyxl — Conditional Updates
# ============================================================

"""### What is a "Conditional Update"?

Imagine you are handed a printed Excel sheet with 500 rows.
Your manager says: "Go through every row. If the Status says
'Pending', change it to 'Under Review'."

Doing that by hand would take an hour.
A Python script does it in under 1 second.

That is what this section is about.

But first — let's create a sample Excel file to practice on.
(We use pandas just to quickly create the file. Don't worry
about pandas yet — that comes later in Section 3.)
"""

import pandas as pd

data = {
    "ID":     ["APP001", "APP002", "APP003", "APP004"],
    "Name":   ["Rahul",  "Priya",  "Sneha",  "Arjun"],
    "Email":  ["rahul@gmail.com", "priya@gmail.com",
               "sneha@gmail.com", "arjun@gmail.com"],
    "Status": ["Pending", "Pending", "Approved", "Pending"]
}

df = pd.DataFrame(data)
df.to_excel("applications.xlsx", index=False, sheet_name="Applications")
print("Sample file created: applications.xlsx")

"""### 1.1 — The Two Modes of iter_rows()

In python_essentials.py you saw `iter_rows()` used like this:

    for row in sheet.iter_rows(min_row=2, values_only=True):
        print(row)   # row = ('APP001', 'Rahul', ...)

That `values_only=True` argument gives you the RAW DATA as a tuple.
It is great for reading, but you CANNOT write back through it.

To WRITE back to a cell, you need the actual Cell object.
A Cell object is like a "remote control" for that specific cell.
You can read its value AND change it.

The difference:
  values_only=True  → gives you plain data  (read only)
  values_only=False → gives you Cell objects (read AND write) ← DEFAULT
"""

from openpyxl import load_workbook

wb    = load_workbook("applications.xlsx")
sheet = wb["Applications"]

print("--- Reading with values_only=True ---")
for row in sheet.iter_rows(min_row=2, values_only=True):
    print(row)   # Plain tuple — you CANNOT change a cell this way

print("\n--- Reading with Cell objects (default) ---")
for row in sheet.iter_rows(min_row=2):
    status_cell = row[3]  # The 4th column (0-indexed) = "Status"
    print(f"Cell {status_cell.coordinate} → value: {status_cell.value}")
    # Now we COULD write: status_cell.value = "something new"

"""### 🚀 Think About It

Look at the second loop above.
- `row` is a tuple of Cell objects, not plain values.
- `row[3]` picks the 4th item (because Python starts counting at 0).
- `.coordinate` tells you the cell address, like "D2".
- `.value` gives you what is inside the cell.

Can you guess what `row[0].value` would print? Try it in the loop above.
"""

"""### 1.2 — Applying the Update

Now that we understand Cell objects, let's do the real automation:
Read each Status cell → If it says "Pending" → Change it → Save.

The important rule: After changing cells, you MUST call `wb.save()`.
Without saving, your changes exist only in memory and are lost.
"""

wb    = load_workbook("applications.xlsx")
sheet = wb["Applications"]

for row in sheet.iter_rows(min_row=2):   # min_row=2 skips the header
    status_cell = row[3]                 # Column D = Status

    if status_cell.value == "Pending":
        status_cell.value = "Under Review"   # ← Changing the cell
        print(f"Row {status_cell.row}: changed to 'Under Review'")

wb.save("applications.xlsx")   # ← Writing changes to the actual file
print("\n✅ All done! Open applications.xlsx to verify.")

"""### 🚀 Your Turn — Challenge 1

Add a brand new column called "Priority" (Column E).

Rules:
  - E1 (the header row) should say "Priority"
  - If a row's Status is "Approved" → set Priority to "High"
  - Everything else               → set Priority to "Normal"

Steps to follow:
  1. Load the workbook
  2. Set sheet["E1"] = "Priority"
  3. Loop through rows starting from row 2
  4. Use row[3].value to read the Status
  5. Use row[4] to get the Priority cell (column E)
  6. Apply your if/else logic
  7. Save the workbook

Write your solution below this docstring and run it!
"""

# Write your solution here ↓
# Done in workbook.py

# ============================================================
# SECTION 2: Cell Formatting
# ============================================================

"""### What is Cell Formatting?

So far we have only changed the DATA inside cells.
But in a real report, you also want it to LOOK good:
  - Header row in bold with a dark background
  - "Approved" rows highlighted in green
  - Important columns in a different color

openpyxl has a `styles` module for exactly this.
You create a style object, then assign it to a cell.

Think of it like this:
  cell.font  = Font(bold=True)     → makes text bold
  cell.fill  = PatternFill(...)    → gives a background color
  cell.alignment = Alignment(...)  → centers the text
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

"""### 2.1 — Styling the Header Row

Colors in openpyxl use HEX codes (the same codes used in web design).
For example:
  "FFFFFF" = White
  "000000" = Black
  "2F4F8F" = Dark Navy Blue
  "C6EFCE" = Light Green

Let's make the header row look like a professional report header.
"""

wb    = load_workbook("applications.xlsx")
sheet = wb["Applications"]

# 1. Define the styles
bold_white   = Font(bold=True, color="FFFFFF")
navy_fill    = PatternFill(start_color="2F4F8F",
                           end_color="2F4F8F",
                           fill_type="solid")
center_text  = Alignment(horizontal="center")

# 2. Apply styles to EVERY cell in row 1
# sheet[1] means "give me all cells in row number 1"
for cell in sheet[1]:
    cell.font      = bold_white
    cell.fill      = navy_fill
    cell.alignment = center_text

print("Header row styled.")

# 3. Now highlight "Approved" rows in light green
green_fill = PatternFill(start_color="C6EFCE",
                         end_color="C6EFCE",
                         fill_type="solid")

for row in sheet.iter_rows(min_row=2):
    if row[3].value == "Approved":
        for cell in row:          # Apply fill to EVERY cell in that row
            cell.fill = green_fill

wb.save("applications.xlsx")
print("✅ Formatting saved. Open the file to see the styled headers!")

"""### 🚀 Your Turn — Challenge 2

Now you try adding two more formatting rules:

1. Rows where Status is "Under Review" should have a YELLOW background.
   (Yellow HEX code: "FFFF00")

2. The Name column (Column B) should be in ITALIC font.
   (Hint: Font(italic=True) — assign it to cell.font)

Load the workbook, loop through rows, apply the rules, save.
"""

# Write your solution here ↓



# ============================================================
# SECTION 3: pandas — Working with Data Like a Pro
# ============================================================

"""### What is pandas and Why Should You Care?

`openpyxl` is great at reading/writing individual cells.
But what if you want to:
  - Find all rows where income > 50000?
  - Calculate the average age of all applicants?
  - Combine data from two different Excel files?

Doing that with openpyxl would take 50 lines of loops and conditions.
With pandas, it takes 1 line.

pandas introduces a new data structure called a **DataFrame**.

A DataFrame is essentially a table — just like an Excel sheet — but
it lives inside Python. You can slice it, filter it, transform it,
and save it back to Excel, all with very clean code.
"""

import pandas as pd

"""### 3.1 — Reading an Excel File

Instead of loading cells one by one with openpyxl,
pandas reads the ENTIRE sheet in one shot.
"""

df = pd.read_excel("applications.xlsx", sheet_name="Applications")

# df is now a DataFrame — let's see what's inside it
print(df)

"""### 3.2 — Inspecting Your Data

Before doing anything with data, you should always inspect it.
These are the four commands every data engineer runs first:
"""

# How many rows and columns?
print("Shape:", df.shape)        # e.g. (4, 5) → 4 rows, 5 columns

# What are the column names?
print("Columns:", df.columns.tolist())

# What does the first 3 rows look like?
print(df.head(3))

# What data type is each column?
print(df.dtypes)

"""### 🚀 Think About It

Run each of those lines one at a time.
What does df.shape return for your applications.xlsx?
What data types does pandas assign to each column?
"""

"""### 3.3 — Selecting a Column

In a dictionary, you access a value with its key: `my_dict["key"]`
In pandas, you access a column the SAME way: `df["ColumnName"]`
"""

# Get the entire "Name" column
print(df["Name"])

# Get the entire "Status" column
print(df["Status"])

# Get MULTIPLE columns at once — pass a LIST of names
print(df[["Name", "Status"]])

"""### 🚀 Your Turn — Challenge 3

1. Print only the "Email" column.
2. Print the "ID" and "Priority" columns together.
"""

# Write your solution here ↓



"""### 3.4 — Filtering Rows

This is the most important pandas skill for automation.

In Excel you would use "AutoFilter" and click dropdowns.
In pandas you describe the condition in code.

Syntax: df[ condition ]

A condition is just a True/False check on a column.
For example: df["Status"] == "Approved"
  → This checks every row and gives back True or False.
  → When you put it inside df[ ], you get back only the True rows.
"""

# Show only approved applications
approved = df[df["Status"] == "Approved"]
print("Approved only:")
print(approved)

# Combining two conditions:
# Use  &  for AND  (both must be true)
# Use  |  for OR   (at least one must be true)
# IMPORTANT: Every condition MUST be inside its own parentheses ()

high_and_approved = df[(df["Status"] == "Approved") & (df["Priority"] == "High")]
print("\nApproved AND High Priority:")
print(high_and_approved)

"""### 🚀 Your Turn — Challenge 4

1. Filter and print all rows where Priority is "Normal".
2. Filter and print rows where Status is "Approved" OR "Under Review".
"""

# Write your solution here ↓



"""### 3.5 — Adding a New Column

Adding a column to a DataFrame is just like adding a new key to a dict.
You give it a name and assign it a value.
"""

# Add a column that extracts the domain from each email
# .str  → tells pandas: "treat this column as text"
# .split("@") → splits "rahul@gmail.com" into ["rahul", "gmail.com"]
# .str[1] → picks the second part: "gmail.com"
df["Domain"] = df["Email"].str.split("@").str[1]

print(df[["Name", "Email", "Domain"]])

"""### 3.6 — Transforming Values with .map()

.map() lets you replace values in a column using a dictionary.
Perfect for adding labels, icons, or display text.
"""

# Map status values to emoji icons
status_icons = {
    "Approved":     "✅",
    "Under Review": "🔄",
    "Pending":      "⏳"
}

df["Status_Icon"] = df["Status"].map(status_icons)
print(df[["Name", "Status", "Status_Icon"]])

"""### 3.7 — Saving a DataFrame to Excel

Once you are done transforming your data,
save it to a new file using .to_excel().

Always save to a NEW file — never overwrite your raw data source!
`index=False` means: don't write the row numbers (0, 1, 2...) as a column.
"""

df.to_excel("applications_processed.xlsx", index=False, sheet_name="Processed")
print("✅ Saved to applications_processed.xlsx")

"""### 🚀 Final Challenge — End-to-End Mini Automation

Put together everything from Section 3:

1. Read applications.xlsx into a DataFrame.
2. Filter only the rows where Status == "Approved".
3. Add a column called "Welcome_Message" with this text for each row:
     "Dear {Name}, your application {ID} has been approved. Welcome!"
   Hint: Use .apply() with a lambda like you saw in python_essentials.py
4. Save those rows ONLY to a new file: approved_applications.xlsx
5. Print the final result to verify.

Try building it step by step — one line at a time.
"""

# Write your solution here ↓


