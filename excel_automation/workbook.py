from openpyxl import load_workbook
import pandas as panda

wb = load_workbook("applications.xlsx")

sheet = wb["Applications"]

sheet["E1"] = "Priority"

for row in sheet.iter_rows(min_row=2):
    if row[3].value == "Approved":
        row[4].value = "High"
    else:
        row[4].value = "Normal"

wb.save("applications.xlsx")
print("\n[SUCCESS] All done! Open applications.xlsx to verify.")

# ------------------------------------------------------------------

df = panda.read_excel("applications.xlsx")

print(df)

print("Columns:", df.columns.tolist())

print(df.head())

print(df["Email"])

print(df[["ID", "Priority"]])

print(df[df["Status"] == "Approved"])
